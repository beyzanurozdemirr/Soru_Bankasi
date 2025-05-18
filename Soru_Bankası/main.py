import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QFileDialog, QMessageBox
from ui.ui_ana_ekran import Ui_MainWindow
from ui.ui_soru_ekleme import Ui_Form as Ui_SoruGirisi
from ui.ui_soru_yazdirma import Ui_Form as Ui_SoruGosterimi
from openpyxl import Workbook, load_workbook
from PyQt5.QtPrintSupport import QPrinter
from PyQt5.QtGui import QTextDocument
import os


class SoruGirisiEkrani(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_SoruGirisi()
        self.ui.setupUi(self)
        self.ui.btnEkle.clicked.connect(self.soru_ekle)
        self.ui.btnKaydet.clicked.connect(self.kaydet)
        self.sorular = []

    def soru_ekle(self):
        soru_metni = self.ui.textSoru.toPlainText()
        cevaplar = [
            self.ui.cevap1.text(),
            self.ui.cevap2.text(),
            self.ui.cevap3.text(),
            self.ui.cevap4.text(),
            self.ui.cevap5.text()
        ]
        dogru_index = -1
        for idx, radio_btn in enumerate([self.ui.radio1, self.ui.radio2, self.ui.radio3, self.ui.radio4, self.ui.radio5]):
            if radio_btn.isChecked():
                dogru_index = idx
                break

        if soru_metni and all(cevaplar) and dogru_index != -1:
            self.sorular.append({
                "soru": soru_metni,
                "cevaplar": cevaplar,
                "dogru": dogru_index
            })
            self.ui.textSoru.clear()
            for cevap_input in [self.ui.cevap1, self.ui.cevap2, self.ui.cevap3, self.ui.cevap4, self.ui.cevap5]:
                cevap_input.clear()
            for radio_btn in [self.ui.radio1, self.ui.radio2, self.ui.radio3, self.ui.radio4, self.ui.radio5]:
                radio_btn.setChecked(False)

    def kaydet(self):
        dosya_adi = "soru_bankasi.xlsx"
        if not self.sorular:
            QMessageBox.warning(self, "Uyarı", "Kaydedilecek soru bulunamadı!")
            return

        if os.path.exists(dosya_adi):
            workbook = load_workbook(dosya_adi)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Soru", "Cevap A", "Cevap B", "Cevap C", "Cevap D", "Cevap E", "Doğru Cevap"])

        for s in self.sorular:
            dogru_harf = chr(65 + s["dogru"])
            sheet.append([s["soru"]] + s["cevaplar"] + [dogru_harf])

        workbook.save(dosya_adi)
        self.sorular.clear()
        QMessageBox.information(self, "Başarılı", "Sorular başarıyla kaydedildi.")


class SoruGosterimiEkrani(QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_SoruGosterimi()
        self.ui.setupUi(self)
        self.ui.btnSec.clicked.connect(self.dosya_sec)
        self.ui.btnYazdir.clicked.connect(self.yazdir)
        self.metin_icerik = ""

    def dosya_sec(self):
        dosya_adi, _ = QFileDialog.getOpenFileName(self, "Dosya Seçiniz", "", "Excel Dosyaları (*.xlsx);;Metin Dosyaları (*.txt)")
        if not dosya_adi:
            return

        if dosya_adi.endswith(".xlsx"):
            workbook = load_workbook(dosya_adi)
            sheet = workbook.active
            metin = ""
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:
                    continue
                soru = row[0]
                cevaplar = row[1:6]
                dogru = row[6]
                metin += f"Soru: {soru}\n"
                for idx, cevap in enumerate(cevaplar):
                    harf = chr(65 + idx)
                    if harf == dogru:
                        metin += f"  {harf}. {cevap} (Doğru)\n"
                    else:
                        metin += f"  {harf}. {cevap}\n"
                metin += "\n"
            self.metin_icerik = metin
            self.ui.textAlan.setPlainText(metin)
        else:
            with open(dosya_adi, "r", encoding="utf-8") as file:
                self.metin_icerik = file.read()
                self.ui.textAlan.setPlainText(self.metin_icerik)

    def yazdir(self):
        dosya_adi, _ = QFileDialog.getSaveFileName(self, "PDF Olarak Kaydet", "", "PDF Dosyası (*.pdf)")
        if not dosya_adi:
            return
        if not dosya_adi.endswith(".pdf"):
            dosya_adi += ".pdf"

        printer = QPrinter(QPrinter.HighResolution)
        printer.setOutputFormat(QPrinter.PdfFormat)
        printer.setOutputFileName(dosya_adi)

        belge = QTextDocument()
        belge.setPlainText(self.metin_icerik)
        belge.print_(printer)


class AnaEkran(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.ui.btnYeniSoru.clicked.connect(self.yeni_soru_ac)
        self.ui.btnSoruSec.clicked.connect(self.soru_sec)

    def yeni_soru_ac(self):
        self.soruGirisi = SoruGirisiEkrani()
        self.soruGirisi.show()

    def soru_sec(self):
        self.soruGosterimi = SoruGosterimiEkrani()
        self.soruGosterimi.show()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    ana_pencere = AnaEkran()
    ana_pencere.show()
    sys.exit(app.exec_())
